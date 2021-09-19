#엑셀 파일 서식 , 스타일 지정하기
import openpyxl
from openpyxl.styles import Font, Alignment , PatternFill, Border , Side

xl = openpyxl.load_workbook("./file/전체광고비.xlsx")
sheet = xl.active

#서식 만들기
font_s = Font(name="맑은 고딕", size=12, bold=True)
ali = Alignment(horizontal="center")
pa = PatternFill(patternType="solid", fgColor="FE9A2E" )
#패턴에 넣는 색은 색깔 코드로 넣어야 한다. #FE9A2E 이런 식으로 나오는 코드이다.
bo = Border(left=Side(style="thin"), right=Side(style="thin") , top=Side(style="thin") , bottom=Side(style="thin"))
#4면을 Border로 얇은 테두리를 만들어 주었다.

sheet["A14"].value = "합계"
sheet["A14"].font = font_s
#이런식으로 하면 입력시 폰트도 적용 가능하다.

sheet["A14"].alignment = ali
#이런식으로 정렬도 가능하다.

sheet["A14"].fill = pa
#이런식으로 색도 넣을 수 있다.

sheet["A14"].border = bo
#이런식으로 테두리도 넣을 수 있다.

sheet["B14"].value = "=SUM(B2:B13)"
#이런식으로 엑셀 안의 함수를 문자열로 입력해서 활용할 수도 있다.

#만약 여러 셀을 지정해서 사용하고 싶다면
#for row in sheet["B14:D14]:
#   for cell in row:
#       cell.font / cell.alignment
#이런식으로 하면 된다.

# 즉 다양한 셀에 접근 하는 방법으로는 이처럼 범위를 정하고 그안의 셀들에 대해 각각 접근하는 방법이 있다.
#만약 B2:D13 하면 B ~ D 와 2 ~ 13 안의 셀들에 대해서 접근하게 되는 것이다.

xl.save("./file/전체광고비.xlsx")